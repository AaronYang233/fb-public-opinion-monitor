#!/usr/bin/env python3
"""
Facebook 舆情报告生成脚本
支持：情感分析、聚类检测、截图嵌入、格式规范输出

用法:
  python3 build_report.py --posts posts.json --output report.xlsx
  python3 build_report.py --posts posts.json --output report.xlsx --screenshots ./screenshots/
  python3 build_report.py --auto   # 自动分析剪贴板中的帖子数据（JSON格式）
"""

import json
import os
import sys
import argparse
import textwrap
import re
from datetime import datetime
from typing import Dict, List, Tuple, Optional

# ========== 中英双语情感关键词库 ==========

STRONG_NEGATIVE_PATTERNS = [
    # 中文强负面
    "退货", "退款", "投诉", "举报", "欺诈", "骗子", "假货", "劣质",
    "质量差", "做工差", "货不对板", "严重", "致命", "完全失败",
    "垃圾", "废物", "毫无用处", "欺骗消费者", "三无产品",
    "有毒", "危险", "安全隐患", "爆炸", "起火", "过热",
    # 英文强负面
    r"\breturn\b.*\b(fraud|scam|ripoff|lawsuit)\b",
    r"\b(fraud|scam|ripoff)\b",
    r"\b(lawsuit|legal action|solicitor)\b",
    r"\bterrible\s*(quality|product|experience)\b",
    r"\bworst\s*car\b",
    r"\bfire\b.*\b(battery|car|charge)\b",
    r"\bexplod\w*\b.*\b(battery|car)\b",
    r"\bpoison\w*\b|\btoxic\b.*\bsmell\b",
    r"\bpermanent\s*damage\b",
    r"\bcan'?t\s*(drive|use|charge)\b.*\b(broken|dead)\b",
    r"\bcompletely\s*(broken|useless|useless)\b",
    r"\b(illegal|dangerous)\s*(product|car)\b",
    r"\bwill\s*never\s*buy\b.*\b(mg|brand)\b",
    r"\bbuyers\s*beware\b",
]

NEGATIVE_PATTERNS = [
    # 中文负面
    "不满", "失望", "差", "糟糕", "难用", "不好", "问题", "故障",
    "损坏", "有瑕疵", "划痕", "变形", "褪色", "异味", "不舒服",
    "延迟", "很久没到", "太慢了", "客服态度差", "推诿",
    "不处理", "没人管", "希望改进", "需要加强",
    # 英文负面
    r"\b(disappointed|unsatisfied|unhappy|angry)\b",
    r"\b(problem|issue|trouble)\s*(with|on|in)\b",
    r"\b(broke|broken|not\s*work)\b.*\b(day|week|month)\b",
    r"\b(fault|faulty|defect)\b",
    r"\b(complaint|complain)\b",
    r"\bnot\s*(good|great|happy|satisfied)\b",
    r"\b(annoying|frustrat\w*)\b",
    r"\b(recall|service\s*center|dealership)\b.*\b(say|said|told)\b.*\b(no|nothing)\b",
    r"\bdealer\s*(didn'?t|won'?t|can'?t)\b",
    r"\bno\s*(solution|fix|answer|help)\b",
    r"\bwaiting\s*(for|ago)\b.*\b(weeks|months)\b",
    r"\bslow\s*(service|response|reply)\b",
    r"\bpoor\s*(quality|service)\b",
    r"\bdoesn'?t\s*(work|function|charge)\b",
    r"\bcan'?t\s*(use|drive|open|start)\b",
    r"\b(won|will).*work.*properly\b",
    r"\b(concern|worried|concerned)\b.*\b(safety|reliability)\b",
    r"\bunnerv\w*\b",
    r"\balarm\b.*\b(false|problem|issue)\b",
    r"\bfalse\s*alarm\b",
]

POSITIVE_PATTERNS = [
    # 中文正面
    "满意", "很好", "优秀", "超值", "推荐", "喜欢", "实用", "物美价廉",
    "性价比高", "客服好", "服务周到", "解决问题", "响应快",
    "好用", "方便", "漂亮", "好看", "正品",
    # 英文正面
    r"\b(love[sd]?|liked|great|excellent|amazing|awesome)\b",
    r"\b(recommend|recommended)\b",
    r"\b(happy|satisfied|pleased|delighted)\b.*\b(with|about)\b",
    r"\b(pleasantly?\s*surprised)\b",
    r"\b(best|good|fair)\s*(value|price|deal)\b",
    r"\bwell\s*(built|designed|made)\b",
    r"\bsmooth\s*(drive|ride)\b",
    r"\bcomfortable\s*(drive|seat|car)\b",
    r"\breal\s*world\s*range\b",
    r"\bbetter\s*than\s*expected\b",
    r"\bno\s*(issue|problem|complaint)\b.*\bso\s*far\b",
    r"\bhappy\s*(with|about)\b.*\b(purchase|car|dealership)\b",
    r"\bworth\s*(every|the)\s*penny\b",
    r"\b(good|great)\s*(experience|dealership)\b",
    r"\b(quick|fast)\s*(service|response)\b",
    r"\bexceed\w*\s*(expectation)\b",
]

QUALITY_PATTERNS = [
    "破损", "瑕疵", "异味", "褪色", "变形", "材质差", "做工粗糙", "假货",
    r"\b(quality|build)\s*(issue|problem|fault)\b",
    r"\bpoor\s*(build\s*quality|material)\b",
    r"\b(broke|broken|scratch|chip|damaged|crease)\b.*\b(day|week|month)\b",
    r"\bfaulty\s*(part|component)\b",
    r"\b(plastic|cheap)\s*(material|feel)\b",
    r"\b(fade[sd]?|discolou?r|warped)\b",
    r"\b(fake|counterfeit)\b",
    r"\b(smell|stink|odor|smells?\s*(of|like))\b",
]

FUNCTION_PATTERNS = [
    "不工作", "故障", "不能用", "功能缺失", "兼容性", "充不进电",
    r"\b(software|system)\s*(bug|issue|problem|update)\b",
    r"\b(screen|display)\s*(not\s*work|blank|glitch)\b",
    r"\bnavigation\s*(issue|problem|doesn'?t\s*work)\b",
    r"\b(key\s*card|fob)\s*(not\s*work|problem)\b",
    r"\b(bluetooth|wifi|connectivity)\s*(issue|problem)\b",
    r"\b(usb|c charging)\s*(not\s*work|problem)\b",
    r"\b(nfc|rfid)\s*(card|key)\s*(issue|problem)\b",
    r"\b(charging|charge)\s*(problem|issue|slow|won'?t\s*charge)\b",
    r"\b(range|anxiety)\s*(issue|problem)\b",
    r"\b(trip\s*planner|navigation)\s*(doesn'?t|didn'?t)\s*plan\b",
    r"\b(alarm)\s*(false|problem|issue|going\s*off)\b",
    r"\b(door|boot|trunk)\s*(won'?t\s*open|open\s*problem)\b",
]

SERVICE_PATTERNS = [
    "退货难", "退款慢", "客服不理", "推诿", "拒绝保修", "维修收费高",
    r"\b(return|refund)\s*(process|difficult|slow)\b",
    r"\b(warranty|guarantee)\s*(denied|rejected|refused)\b",
    r"\b(service\s*center|dealership)\s*((no|nothing)\s*they\s*can\s*do)\b",
    r"\bcustomer\s*service\s*(poor|bad|terrible|useless)\b",
    r"\b(dealer|dealership)\s*(ignored|didn'?t\s*help)\b",
    r"\brepair\s*(cost|price|expensive|overpriced)\b",
    r"\b(no\s*response|unresponsive|ignored)\b",
    r"\b(complaint|escalate[sd])\b",
    r"\b(legal|lawsuit|solicitor)\b",
]

SAFETY_PATTERNS = [
    "漏电", "起火", "爆炸", "有毒", "过热", "夹手", "锐利边缘",
    r"\b(fire|flame|burn)\b.*\b(battery|car|charging)\b",
    r"\b(smoke|burning\s*smell)\b.*\b(car|battery|charge)\b",
    r"\b(explod\w*|explosion)\b",
    r"\b(overheat\w*|overheating)\b.*\b(battery|car|charge)\b",
    r"\b(safety|danger)\s*(concern|issue|recall)\b",
    r"\b(recall)\b",
    r"\b(poison\w*|toxic)\s*(smell|gas|fume)\b",
    r"\b(life\s*threat|unsafe|dangerous)\b",
    r"\b(brake\s*fail|steering\s*fail|accelerat\w*)\s*(issue|problem)\b",
]


def _match_any(text: str, patterns: List[str]) -> int:
    """返回文本匹配到的模式数量"""
    text_lower = text.lower()
    count = 0
    for p in patterns:
        if re.search(p, text_lower):
            count += 1
    return count


def analyze_sentiment(text: str) -> Tuple[str, int]:
    """双语情感分析"""
    # 强负面优先
    strong_count = _match_any(text, STRONG_NEGATIVE_PATTERNS)
    if strong_count >= 2:
        return "强负面", 4
    if strong_count >= 1:
        return "强负面", 4

    neg_count = _match_any(text, NEGATIVE_PATTERNS)
    if neg_count >= 3:
        return "负面", 3
    if neg_count >= 1:
        return "负面", 3

    pos_count = _match_any(text, POSITIVE_PATTERNS)
    if pos_count >= 2:
        return "正面", 1
    if pos_count >= 1:
        return "正面", 1

    return "中性", 2


def classify_issue(text: str) -> str:
    """双语问题分类（按优先级）"""
    if _match_any(text, SAFETY_PATTERNS) > 0:
        return "安全风险"
    if _match_any(text, QUALITY_PATTERNS) > 0:
        return "质量问题"
    # 功能异常：false alarm / alarm malfunction 独立检测（优先级高于售后）
    if re.search(r"false\s*alarm|alarm\s*(going\s*off|malfunction|problem|issue)", text.lower()):
        return "功能异常"
    if _match_any(text, FUNCTION_PATTERNS) > 0:
        return "功能异常"
    if _match_any(text, SERVICE_PATTERNS) > 0:
        return "售后问题"
    return "其他"


def assess_severity(
    sentiment: str,
    sentiment_score: int,
    likes: int,
    comments: int,
    issue_type: str,
    user_influence: int = 2,
) -> Tuple[str, float]:
    """严重程度评估"""
    emotion_score = sentiment_score

    total_engagement = likes + comments
    if total_engagement >= 100:
        spread_score = 4
    elif total_engagement >= 50:
        spread_score = 3
    elif total_engagement >= 10:
        spread_score = 2
    else:
        spread_score = 1

    type_scores = {
        "安全风险": 4, "质量问题": 3,
        "功能异常": 3, "售后问题": 2, "其他": 1
    }
    type_score = type_scores.get(issue_type, 1)

    # 特殊加权：强负面+高互动 → 立即升级
    if sentiment == "强负面" and total_engagement >= 20:
        type_score = max(type_score, 4)
        spread_score = max(spread_score, 3)

    total_score = (
        emotion_score * 0.3
        + spread_score * 0.3
        + type_score * 0.25
        + user_influence * 0.15
    )

    if total_score >= 3.5:
        return "严重", total_score
    elif total_score >= 2.5:
        return "高", total_score
    elif total_score >= 1.5:
        return "中", total_score
    else:
        return "低", total_score


def assess_spread_risk(likes: int, comments: int) -> str:
    """扩散风险评估"""
    total = likes + comments
    if total >= 100:
        return "高"
    elif total >= 50:
        return "中"
    return "低"


def generate_suggestion(issue_type: str, sentiment: str) -> str:
    """生成双语处理建议"""
    suggestions = {
        "质量问题": "建议联系售后换货，保留购买凭证，如多次投诉建议抽检同批次产品",
        "功能异常": "建议提交工单技术支持，必要时申请退换货，排查是否为个例或系统缺陷",
        "售后问题": "建议升级投诉级别，联系总部客服，记录工单编号跟进",
        "安全风险": "⚠️ 立即下架相关批次产品，进行安全检测，发布官方声明",
        "其他": "建议收集更多案例，分析是否为个案或系统性问题",
    }
    base = suggestions.get(issue_type, "常规处理")

    # 强负面 → 加急措辞
    if sentiment == "强负面" and issue_type not in ["安全风险"]:
        base = "⚠️ " + base

    return base


# ========== Excel 报告生成 ==========

def build_report(
    posts: List[Dict],
    output_path: str,
    screenshot_dir: str = "",
    group_name: str = "",
    scan_date: str = "",
) -> str:
    """
    生成格式规范的舆情分析 Excel 报告

    返回: 报告文件路径
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ 缺少 openpyxl，运行: pip install openpyxl")
        sys.exit(1)

    # ── 样式定义 ──────────────────────────────────────
    HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # 深蓝
    ALERT_FILL    = PatternFill("solid", fgColor="FF0000")   # 红色预警
    WARN_FILL     = PatternFill("solid", fgColor="FF7F00")  # 橙色警告
    GOOD_FILL     = PatternFill("solid", fgColor="70AD47")  # 绿色正面
    ROW_FILL_A    = PatternFill("solid", fgColor="FFFFFF")  # 白色行
    ROW_FILL_B    = PatternFill("solid", fgColor="F2F2F2")  # 浅灰行
    BORDER_COLOR  = "BFBFBF"

    HEADER_FONT   = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    BOLD_FONT     = Font(name="微软雅黑", bold=True, size=10)
    NORMAL_FONT   = Font(name="微软雅黑", size=10)
    SMALL_FONT    = Font(name="微软雅黑", size=9)
    LINK_FONT     = Font(name="微软雅黑", size=9, color="0563C1", underline="single")

    thin = Side(style="thin", color=BORDER_COLOR)
    THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    RIGHT  = Alignment(horizontal="right",  vertical="center")

    # ── 工作簿（仅一个 Sheet）──────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "舆情分析"
    ws.sheet_view.showGridLines = False

    # 表头行（固定）
    headers = [
        ("序号",       6),
        ("Complain",  46),   # 截图图片
        ("Like",       7),
        ("Comments",   10),
        ("Remark",    58),   # 原文 + 总结
        ("Link",      40),
        ("粉丝评论",  40),
    ]
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # 填充数据行
    for idx, post in enumerate(posts, start=1):
        row_num = idx + 1
        is_alt  = idx % 2 == 0
        row_fill = ROW_FILL_B if is_alt else ROW_FILL_A

        sentiment = post.get("sentiment", "中性")
        severity  = post.get("severity", "低")
        issue_type = post.get("issue_type", "其他")

        # Remark 内容（简化格式：原文 + 总结）
        original_content = post.get("content", "")
        # 生成简短总结
        summary_parts = []
        if sentiment != "中性":
            summary_parts.append(f"情感倾向：{sentiment}")
        if issue_type != "其他":
            summary_parts.append(f"问题类型：{issue_type}")
        if severity != "低":
            summary_parts.append(f"严重程度：{severity}")
        
        summary_text = "；".join(summary_parts) if summary_parts else "内容正常"
        
        remark_text = f"原文：{original_content}\n\n总结内容：{summary_text}"

        row_values = [
            idx,                                    # 序号
            "" ,
            post.get("likes", 0),                   # Like
            post.get("comments", 0),                # Comments
            remark_text,                            # Remark
            post.get("link", ""),                   # Link
            post.get("fan_comments", ""),          # 粉丝评论
        ]

        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border  = THIN_BORDER
            cell.fill    = row_fill
            cell.font    = SMALL_FONT if col_idx not in [1, 2, 5] else NORMAL_FONT

            if col_idx == 1:        # 序号 → 居中
                cell.alignment = CENTER
            elif col_idx == 3:      # Like → 居中
                cell.alignment = CENTER
            elif col_idx == 4:      # Comments → 居中
                cell.alignment = CENTER
            elif col_idx == 5:      # Remark → 左上
                cell.alignment = LEFT
            elif col_idx == 6:      # Link → 超链接字体
                cell.font = LINK_FONT
                cell.alignment = LEFT
            else:
                cell.alignment = LEFT

            # 序号列按严重程度着色
            if col_idx == 1:
                if severity in ["严重"]:
                    cell.fill = ALERT_FILL
                    cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
                elif severity == "高":
                    cell.fill = WARN_FILL
                    cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
                elif sentiment == "正面":
                    cell.fill = GOOD_FILL
                    cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)

        ws.row_dimensions[row_num].height = 80

        # 嵌入截图（优先从截图目录匹配）
        post_id = post.get("post_id", "")
        if screenshot_dir and os.path.isdir(screenshot_dir):
            matched = [
                f for f in os.listdir(screenshot_dir)
                if f.startswith(str(post_id))
                and f.lower().endswith((".png", ".jpg", ".jpeg"))
            ]
            if matched:
                img_path = os.path.join(screenshot_dir, matched[0])
                try:
                    img = XLImage(img_path)
                    img.width  = 420
                    img.height = 260
                    ws.add_image(img, f"B{row_num}")
                except Exception as e:
                    # 截图失败 → B列保持为空，不写文本
                    print(f"  ⚠️ 截图嵌入失败 [{post_id}]: {e}")

    # 保存
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


# ========== CLI 入口 ==========

def main():
    parser = argparse.ArgumentParser(description="生成 FB 舆情分析 Excel 报告")
    parser.add_argument("--posts",   help="帖子数据 JSON 文件路径")
    parser.add_argument("--output",  help="输出 Excel 文件路径", required=True)
    parser.add_argument("--screenshots", help="截图目录路径", default="")
    parser.add_argument("--group",   help="群组名称", default="")
    parser.add_argument("--date",    help="扫描日期 YYYY-MM-DD", default="")
    parser.add_argument("--auto",    help="从 stdin 读取 JSON", action="store_true")
    args = parser.parse_args()

    # 加载帖子数据
    if args.auto:
        posts = json.load(sys.stdin)
    elif args.posts:
        with open(args.posts) as f:
            posts = json.load(f)
    else:
        print("❌ 请指定 --posts <文件> 或使用 --auto")
        sys.exit(1)

    # 对每条帖子运行分析
    for post in posts:
        text = post.get("content", "")
        sentiment, score = analyze_sentiment(text)
        issue_type = classify_issue(text)
        severity, sscore = assess_severity(
            sentiment, score,
            post.get("likes", 0),
            post.get("comments", 0),
            issue_type,
        )
        spread_risk = assess_spread_risk(post.get("likes", 0), post.get("comments", 0))
        suggestion  = generate_suggestion(issue_type, sentiment)

        post["sentiment"]    = sentiment
        post["severity"]     = severity
        post["issue_type"]   = issue_type
        post["spread_risk"]  = spread_risk
        post["suggestion"]   = suggestion

    output = build_report(
        posts=posts,
        output_path=args.output,
        screenshot_dir=args.screenshots,
        group_name=args.group,
        scan_date=args.date,
    )

    print(f"✅ 报告已生成: {output}")


if __name__ == "__main__":
    main()
